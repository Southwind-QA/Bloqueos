# -*- coding: utf-8 -*-
"""Cruce de bloqueos: stock de todas las bodegas x LAB-REG-08 x registro de detenciones.

Dos origenes de bloqueo con naturaleza distinta:
  - LAB        : derivado. Se recalcula entero en cada corrida.
  - DETENCION  : declarado por correo. Entrada durable, el script solo la lee.
Se acumulan: un lote con ambos necesita cerrar los dos.
"""
import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

BASE = config.BASE
LABS = sorted(glob.glob(config.ruta(config.GLOB_LAB)))
F_DET = config.ruta(config.ARCH_DETENCIONES)
OUT = config.ruta(config.ARCH_SALIDA)

LIM_RAM = config.LIM_RAM
LIM_NITRITO = config.LIM_NITRITO
MIN_LOTE = config.MIN_LOTE

LM = ["LM1", "LM2", "LM3", "LM4", "LM5"]
RAM = ["RAM1", "RAM2", "RAM3", "RAM4", "RAM5"]
NIT = ["NITRITO", "NITRITO.1", "NITRITO.2"]


def norm(s):
    """Lote comparable. Conserva @ y B: son parte del codigo, no ruido.
    2AS2621170M y B2AS2621170M son lotes DISTINTOS (correo de detencion 09/07/2026).
    Solo se corta el sufijo *SEMANA y se limpian separadores."""
    return re.sub(r"[^A-Z0-9@]", "", str(s).strip().upper().split("*")[0])


def laxo(s):
    """Version tolerante, solo para detectar posibles errores de tipeo en @ / B."""
    return norm(s).lstrip("@").lstrip("B")


def emparenta(a, b):
    """True si a y b son el mismo lote a distinto nivel de detalle (lote vs sublote)."""
    if len(a) < MIN_LOTE or len(b) < MIN_LOTE:
        return a == b
    return a.startswith(b) or b.startswith(a)


# ----------------------------------------------------------------- STOCK (N bodegas)
marcos = []
for f in sorted(glob.glob(os.path.join(BASE, "*.xlsx"))):
    n = os.path.basename(f)
    if n.startswith("~$") or n.startswith("LAB-REG-08") or f in (F_DET, OUT):
        continue
    try:
        cab = pd.read_excel(f, header=None, nrows=2)
        if "LOTE DE PLANTA" not in cab.iloc[1].astype(str).values:
            print("  (omitido, no parece stock):", n)
            continue
        d = pd.read_excel(f, header=1).dropna(how="all")
    except Exception as e:                                    # noqa: BLE001
        print("  (omitido, no se pudo leer):", n, e)
        continue
    titulo = str(cab.iloc[0, 0])
    d["BODEGA"] = titulo.split("-")[-1].strip() if "-" in titulo else titulo.strip()
    d["ARCHIVO ORIGEN"] = n
    marcos.append(d)
    print(f"  stock: {n}  ->  {len(d)} cajas  |  bodega: {d['BODEGA'].iloc[0]}")

if not marcos:
    raise SystemExit("No se encontro ningun archivo de stock.")
stock = pd.concat(marcos, ignore_index=True)
stock = stock[stock["LOTE DE PLANTA"].notna()].copy()
stock["_L"] = stock["LOTE DE PLANTA"].map(norm)

# ----------------------------------------------------------------- LAB (N anios)
# Los archivos por anio no son identicos: 2025 escribe "Nitrito" y no trae la columna
# "A/R Nitrito"; 2026 escribe "NITRITO" y si la trae. Se ubica cada bloque por nombre,
# no por posicion.
INFO = ["CÓDIGO LAB", "FECHA INGRESO", "TIPO ", "GRUPO", "PRESENTACIÓN", "LOTE SW", "OBSERVACIÓN"]


def carga_lab(f):
    n = os.path.basename(f)
    xls = pd.ExcelFile(f)
    hojas = [h for h in xls.sheet_names if h.strip().upper().startswith("RESULTADOS MICRO")]
    if not hojas:
        print("  (lab omitido, sin hoja RESULTADOS MICRO):", n)
        return None
    d = pd.read_excel(f, sheet_name=hojas[0], header=2).dropna(how="all")
    d = d[d["LOTE SW"].notna()].copy()
    cols = list(d.columns)

    nit = [c for c in cols if re.fullmatch(r"NITRITO(\.\d+)?", str(c).strip().upper())]
    tope = max(cols.index(c) for c in nit) if nit else -1
    prom = next((c for c in cols
                 if str(c).strip().upper().startswith("PROMEDIO") and cols.index(c) > tope), None)
    ar = next((c for c in cols if "A/R" in str(c).upper()), None)

    out = pd.DataFrame(index=d.index)
    for c in INFO:
        out[c] = d[c] if c in d.columns else None
    for i, c in enumerate(LM):
        out[c] = d[c] if c in d.columns else None
    for c in RAM:
        out[c] = d[c] if c in d.columns else None
    for i, c in enumerate(NIT):
        out[c] = d[nit[i]] if i < len(nit) else None
    out["NITRITO PROMEDIO"] = d[prom] if prom else None
    out["A/R Nitrito"] = d[ar] if ar else ""
    out["FUENTE LAB"] = hojas[0]
    print(f"  lab: {n}  ->  {len(out)} muestras con lote  |  hoja: {hojas[0]}"
          + ("" if ar else "  (sin columna A/R Nitrito)"))
    return out


marcos_lab = [x for x in (carga_lab(f) for f in LABS) if x is not None]
if not marcos_lab:
    raise SystemExit("No se encontro ningun LAB-REG-08 legible.")
lab = pd.concat(marcos_lab, ignore_index=True)

lab["_L"] = lab["LOTE SW"].map(norm)
lab["_FECHA"] = pd.to_datetime(lab["FECHA INGRESO"], errors="coerce")
_lm = lab[LM].astype(str).apply(lambda c: c.str.strip().str.upper())
lab["_LM_P"] = (_lm == "P").any(axis=1)
lab["_LM_DATO"] = lab[LM].notna().any(axis=1)
lab["_RAM_MAX"] = lab[RAM].apply(pd.to_numeric, errors="coerce").max(axis=1)
_n = lab[NIT].apply(pd.to_numeric, errors="coerce")
lab["_NIT"] = pd.to_numeric(lab["NITRITO PROMEDIO"], errors="coerce").fillna(_n.mean(axis=1))
lab["_NIT_MIN"] = _n.min(axis=1)
lab["_AR"] = lab["A/R Nitrito"].astype(str).str.strip().str.upper().replace("NAN", "")

# ------------------------------------------------- linea de proceso
# El nitrito es un control de la linea REFRIGERADA: en la congelada el criterio no
# aplica. La clasificacion sale primero del texto del laboratorio, que es el que
# nombra la linea, y si no alcanza, de la condicion del producto en bodega
# (CONDICION concuerda con el nombre en todos los productos que lo declaran).
# Ante la duda NO se clasifica como congelada: ignorar un criterio es la direccion
# permisiva y no se toma por defecto.
_COND = "CONDICIÓN" if "CONDICIÓN" in stock.columns else "CONDICION"
_porlote = stock.groupby("_L").agg(
    cond=(_COND, lambda x: set(x.dropna().astype(str).str.upper())),
    nom=("NOMBRE PRODUCTO", lambda x: " ".join(set(x.dropna().astype(str))).lower()))
_LOTES_ST = sorted(_porlote.index)


def linea_stock(l):
    """Linea de un lote segun su stock. Vacio = no se pudo determinar."""
    if l not in _porlote.index:
        return ""
    r = _porlote.loc[l]
    if "refrigerad" in r["nom"] or ({"REFRIGERADO", "FRESCO"} & r["cond"]):
        return "REFRIGERADA"
    if "carpaccio" in r["nom"] or r["cond"] == {"CONGELADO"}:
        return "CONGELADA"
    return ""


_txt = (lab["GRUPO"].fillna("").astype(str) + " " + lab["PRESENTACIÓN"].fillna("").astype(str)
        + " " + lab["OBSERVACIÓN"].fillna("").astype(str)).str.lower()


def _linea_muestra(i, l):
    t = _txt.iloc[i]
    if "refrigerad" in t:
        return "REFRIGERADA"
    if "carpaccio" in t or "congelad" in t:
        return "CONGELADA"
    base = next((x for x in _LOTES_ST if len(x) >= MIN_LOTE
                 and (l.startswith(x) or x.startswith(l))), None)
    return linea_stock(base) if base else ""


# ---- destino EE.UU.
# Listeria es permisible en linea congelada, salvo lo que se exporta a EE.UU.
# Se usan dos señales: el cliente del stock (dato directo del destino) y el texto
# del producto (wheel, bacon, nombre en ingles). Basta una para marcar destino USA.
# Un lote con varios clientes queda marcado si CUALQUIERA es de EE.UU.: las cajas
# comparten lote y no se pueden separar.
US_CLI = re.compile(config.US_CLI, re.I)
CR_CLI = re.compile(config.CR_CLI, re.I)
# Bacon y wheel salen congelados de planta pero se venden refrigerados en destino:
# para el nitrito se evaluan como linea refrigerada, no como congelada.
BACON_WHEEL = re.compile(config.BACON_WHEEL, re.I)
US_PROD = re.compile(config.US_PROD, re.I)
_cli = stock.groupby("_L").agg(
    cli=("CLIENTE", lambda x: " | ".join(sorted(set(x.dropna().astype(str))))),
    nom=("NOMBRE PRODUCTO", lambda x: " | ".join(sorted(set(x.dropna().astype(str))))))


def destino_restringido(l):
    """Destino que exige ausencia de Listeria aunque el producto sea congelado.

    Devuelve "EE.UU.", "Costa Rica", "" (ningun destino restringido) o None si no
    hay con que determinarlo. None no equivale a "": sin informacion el criterio
    se aplica igual.
    """
    if l is None or l not in _cli.index:
        return None
    r = _cli.loc[l]
    if CR_CLI.search(r["cli"]):
        return "Costa Rica"
    if US_CLI.search(r["cli"]) or US_PROD.search(r["nom"]):
        return "EE.UU."
    return ""


lab["_LINEA"] = [_linea_muestra(i, l) for i, l in enumerate(lab["_L"])]


def _bw(i, l):
    """Bacon o wheel, segun el texto del laboratorio o el producto en bodega."""
    if BACON_WHEEL.search(_txt.iloc[i]):
        return True
    base = next((x for x in _LOTES_ST if len(x) >= MIN_LOTE
                 and (l.startswith(x) or x.startswith(l))), None)
    return bool(base and BACON_WHEEL.search(_cli.loc[base, "nom"]))


lab["_BW"] = [_bw(i, l) for i, l in enumerate(lab["_L"])]
# El nitrito aplica a la linea refrigerada y a bacon/wheel aunque figuren congelados.
lab["_NIT_APLICA"] = (lab["_LINEA"] != "CONGELADA") | lab["_BW"]
print(f"  bacon/wheel: {int(lab['_BW'].sum())} muestras se evaluan como refrigeradas "
      "para el nitrito")


def _restr_muestra(i, l):
    if US_PROD.search(_txt.iloc[i]):
        return "EE.UU."
    base = next((x for x in _LOTES_ST if len(x) >= MIN_LOTE
                 and (l.startswith(x) or x.startswith(l))), None)
    return destino_restringido(base)


# Listeria aplica siempre a linea refrigerada. En congelada solo si el destino es
# EE.UU. o Costa Rica, o si no se pudo determinar: no se exime un criterio por
# falta de informacion.
_restr = [_restr_muestra(i, l) for i, l in enumerate(lab["_L"])]
lab["_DESTINO"] = [x if x is not None else "" for x in _restr]
lab["_LIS_APLICA"] = [(ln != "CONGELADA") or (d is None) or (d != "")
                      for ln, d in zip(lab["_LINEA"], _restr)]
_ex = lab[lab["_LM_DATO"] & ~lab["_LIS_APLICA"]]
print(f"  listeria: no aplica a {len(_ex)} muestras de linea congelada sin destino "
      f"restringido (de {int(lab['_LM_DATO'].sum())} con dato de listeria)")
_c = lab.loc[lab["_NIT"].notna(), "_LINEA"].replace("", "sin clasificar").value_counts()
print("  linea de proceso en muestras con nitrito: "
      + ", ".join(f"{k}={v}" for k, v in _c.items()))

# ----------------------------------------------------------------- DETENCIONES
det = pd.read_excel(F_DET, sheet_name="DETENCIONES", header=4).dropna(how="all")
det = det[det["ID"].notna()].copy()
det["_L"] = det["LOTE"].map(norm)
det["_EVENTO"] = pd.to_datetime(det["FECHA DEL EVENTO"], errors="coerce")
det["_ESTADO"] = det["ESTADO"].astype(str).str.strip().str.upper()
det["_VIGENTE"] = det["_ESTADO"].isin(["ABIERTA", "PNC"])


def evalua(rows, criterios=None, desde=None):
    """Estado de cada criterio sobre un conjunto de muestras de laboratorio."""
    if desde is not None:
        rows = rows[rows["_FECHA"].notna() & (rows["_FECHA"] >= desde)]
    r = {"n": len(rows)}
    if len(rows) == 0:
        r.update(listeria=None, ram=None, nitrito=None, nit_min=None, ar="")
        return r, rows
    r["listeria"] = ("PRESENCIA" if rows["_LM_P"].any()
                     else ("Ausencia" if rows["_LM_DATO"].any() else None))
    r["ram"] = rows["_RAM_MAX"].max() if rows["_RAM_MAX"].notna().any() else None
    r["nitrito"] = rows["_NIT"].min() if rows["_NIT"].notna().any() else None
    r["nit_min"] = rows["_NIT_MIN"].min() if rows["_NIT_MIN"].notna().any() else None
    r["ar"] = "R" if (rows["_AR"] == "R").any() else ("A" if (rows["_AR"] == "A").any() else "")
    return r, rows


def incumple(r, criterio):
    """(no_conforme, falta_dato, texto) para un criterio dado."""
    if criterio == "LISTERIA":
        if r["listeria"] is None:
            return False, True, "sin dato de listeria"
        return r["listeria"] == "PRESENCIA", False, "Listeria: PRESENCIA"
    if criterio == "RAM":
        if r["ram"] is None:
            return False, True, "sin dato de RAM"
        return r["ram"] > LIM_RAM, False, f"RAM {r['ram']:,.0f} UFC/g > {LIM_RAM:,}".replace(",", ".")
    if criterio == "NITRITO":
        if r["nitrito"] is None:
            return False, True, "sin dato de nitrito"
        return r["nitrito"] < LIM_NITRITO, False, f"Nitrito {r['nitrito']:.1f} ppm < {LIM_NITRITO}"
    return False, True, f"criterio no reconocido: {criterio}"


# ------------------------------------------------- LIBERACIONES DECLARADAS
# Bloqueo 2026.xlsm es el registro operativo. La hoja Historia es el log de
# liberaciones con su mercado; la hoja principal marca Estado='liberado'.
# Se toman SOLO los registros explicitos de liberacion: las filas sin estado
# quedan fuera hasta definir que significan, porque interpretarlas mal esconde
# bloqueos o inventa liberaciones.
F_OPE = config.ruta(config.ARCH_OPERATIVO)
MERCADOS = ("Nacional", "Exportación", "USA")
reg = []
if os.path.exists(F_OPE):
    try:
        _h = pd.read_excel(F_OPE, sheet_name="Historia")
        _h.columns = [str(c).strip() for c in _h.columns]
        for _, x in _h[_h["Lote"].notna()].iterrows():
            mk = [m for m in MERCADOS
                  if str(x.get(m, "")).strip().lower() in ("si", "sí", "x", "ok", "1")]
            reg.append({"_L": norm(x["Lote"]), "FECHA": pd.to_datetime(x["Fecha"], errors="coerce"),
                        "MERCADOS": "/".join(mk), "FUENTE": "Historia"})
    except Exception as e:                                              # noqa: BLE001
        print("  (no se pudo leer la hoja Historia):", e)
    try:
        _o = pd.read_excel(F_OPE, sheet_name="codigos bloqueados SAP", header=1).dropna(how="all")
        _o = _o[_o["Traza"].notna()]
        _o = _o[_o["Estado"].astype(str).str.strip().str.lower() == "liberado"]
        for _, x in _o.iterrows():
            mk = [m for m, c in zip(MERCADOS, ("NACIONAL", "EXPORTACIÓN", "USA"))
                  if pd.notna(x.get(c))]
            reg.append({"_L": norm(x["Traza"]),
                        "FECHA": pd.to_datetime(x["Fecha Liberacion"], errors="coerce"),
                        "MERCADOS": "/".join(mk), "FUENTE": "Registro SAP"})
    except Exception as e:                                              # noqa: BLE001
        print("  (no se pudo leer la hoja de codigos bloqueados):", e)
else:
    print("  (sin Bloqueo 2026.xlsm: no hay liberaciones declaradas que cruzar)")

if reg:
    _lib = pd.DataFrame(reg)
    _lib = _lib[_lib["_L"].str.len() >= MIN_LOTE]
    libu = _lib.groupby("_L").agg(
        FECHA=("FECHA", "max"),
        MERCADOS=("MERCADOS", lambda x: "/".join(sorted({k for v in x for k in str(v).split("/") if k}))),
        FUENTE=("FUENTE", lambda x: " / ".join(sorted(set(x)))))
    print(f"  liberaciones declaradas: {len(_lib)} registros sobre {len(libu)} lotes"
          + (f", hasta {_lib['FECHA'].max():%d/%m/%Y}" if _lib["FECHA"].notna().any() else ""))
else:
    libu = pd.DataFrame(columns=["FECHA", "MERCADOS", "FUENTE"])
LIBK = list(libu.index)


def liberacion(clave):
    """Liberacion declarada para un lote o batch: fecha, mercados y de donde sale."""
    k = [x for x in LIBK if emparenta(x, clave)]
    if not k:
        return "", "", ""
    m = libu.loc[k]
    f = m["FECHA"].max()
    mk = sorted({v for x in m["MERCADOS"] for v in str(x).split("/") if v})
    return (f.strftime("%Y-%m-%d") if pd.notna(f) else ""), "/".join(mk), \
        " / ".join(sorted({v for x in m["FUENTE"] for v in str(x).split(" / ")}))


CRIT = ("LISTERIA", "RAM", "NITRITO")


def cod_lab(v):
    """El codigo de laboratorio a veces viene como numero (9250.0) y a veces con
    letra ('1623A'). Un rstrip aqui se come los ceros finales."""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else str(v).strip()
    except (TypeError, ValueError):
        return str(v).strip() or "?"


def con_dato(rows, c):
    """Muestras que midieron el criterio, aplique o no."""
    if c == "LISTERIA":
        return rows[rows["_LM_DATO"]]
    if c == "RAM":
        return rows[rows["_RAM_MAX"].notna()]
    return rows[rows["_NIT"].notna()]


def mide(rows, c):
    """Muestras que midieron el criterio Y a las que el criterio aplica.

    RAM aplica a todo. Nitrito solo a linea refrigerada. Listeria a refrigerada y a
    congelada con destino EE.UU. o Costa Rica. Una muestra sin clasificar NO se
    excluye: no se exime un criterio por falta de informacion.
    """
    if c == "LISTERIA":
        return rows[rows["_LM_DATO"] & rows["_LIS_APLICA"]]
    if c == "RAM":
        return rows[rows["_RAM_MAX"].notna()]
    return rows[rows["_NIT"].notna() & rows["_NIT_APLICA"]]


def por_que_no_aplica(rows, c):
    """Texto trazable para un criterio medido pero no exigible, con el valor omitido."""
    if c == "NITRITO":
        v = rows["_NIT"].min()
        return ("Nitrito no aplica: linea congelada (no es bacon ni wheel)"
                + (f" (se omite {v:.1f} ppm, bajo {LIM_NITRITO})" if pd.notna(v)
                   and v < LIM_NITRITO else ""))
    if c == "LISTERIA":
        hubo = bool(rows["_LM_P"].any())
        return ("Listeria no aplica: linea congelada sin destino EE.UU. ni Costa Rica"
                + (" (se omite una PRESENCIA)" if hubo else ""))
    return f"{c.title()} no aplica"


def _falla(row, c):
    if c == "LISTERIA":
        return bool(row["_LM_P"])
    if c == "RAM":
        return pd.notna(row["_RAM_MAX"]) and row["_RAM_MAX"] > LIM_RAM
    return pd.notna(row["_NIT"]) and row["_NIT"] < LIM_NITRITO


def _texto(row, c):
    if c == "LISTERIA":
        return "Listeria: PRESENCIA"
    if c == "RAM":
        return f"RAM {row['_RAM_MAX']:,.0f} UFC/g > {LIM_RAM:,}".replace(",", ".")
    return f"Nitrito {row['_NIT']:.1f} ppm < {LIM_NITRITO}"


def historia(rows, c):
    """Estado VIGENTE de un criterio, no el peor de toda su historia.

    Un resultado reprobado deja de estar vigente solo si despues hay muestras que
    vuelven a medir ESE criterio y salen conformes. El flujo real en planta es
    fallar, tratar (tipicamente altas presiones) y re-muestrear; una muestra
    posterior que no midio lo que fallo no es evidencia de nada, y por eso se
    filtra con mide() antes de comparar fechas.

    Un re-muestreo conforme NO libera por si solo: deja el lote como candidato y
    la decision la firma Calidad.
    """
    m = mide(rows, c).sort_values("_FECHA")
    vacio = {"estado": "SIN DATO", "txt": "", "falla": pd.NaT, "ok": pd.NaT, "n_post": 0}
    if m.empty:
        # medido pero no exigible: se distingue de "nadie lo midio", porque la razon
        # de la liberacion es distinta y hay que poder rastrearla
        td = con_dato(rows, c)
        if len(td):
            return {**vacio, "estado": "NO APLICA", "txt": por_que_no_aplica(td, c)}
        return vacio
    malo = m.apply(lambda r: _falla(r, c), axis=1)
    if not malo.any():
        return {**vacio, "estado": "CONFORME"}
    ult = m[malo].iloc[-1]
    post = m[m["_FECHA"] > ult["_FECHA"]]
    if len(post) and not post.apply(lambda r: _falla(r, c), axis=1).any():
        obs = " / ".join(sorted({str(v).strip() for v in post["OBSERVACIÓN"].dropna()}))[:60]
        return {"estado": "REMUESTREO CONFORME",
                "txt": (f"{_texto(ult, c)} el {ult['_FECHA']:%d/%m/%Y}, luego {len(post)} "
                        f"muestra(s) conforme(s) hasta el {post['_FECHA'].max():%d/%m/%Y}"
                        + (f" [{obs}]" if obs else "")),
                "falla": ult["_FECHA"], "ok": post["_FECHA"].max(), "n_post": len(post)}
    return {"estado": "NO CONFORME", "txt": _texto(ult, c),
            "falla": ult["_FECHA"], "ok": pd.NaT, "n_post": len(post)}


def resume(rows):
    """Estado por criterio de un conjunto de muestras, agregando sus batches."""
    if len(rows) == 0:
        return {c: {"estado": "SIN DATO", "txt": ""} for c in CRIT}, {}
    porbat = {b: {c: historia(g, c) for c in CRIT} for b, g in rows.groupby("_L")}
    agg = {}
    for c in CRIT:
        e = [porbat[b][c] for b in porbat]
        nc = [x for x in e if x["estado"] == "NO CONFORME"]
        rc = [x for x in e if x["estado"] == "REMUESTREO CONFORME"]
        if nc:
            agg[c] = {"estado": "NO CONFORME", "txt": "; ".join(sorted({x["txt"] for x in nc}))}
        elif rc:
            agg[c] = {"estado": "REMUESTREO CONFORME",
                      "txt": "; ".join(sorted({x["txt"] for x in rc}))}
        elif any(x["estado"] == "CONFORME" for x in e):
            n = sum(1 for x in e if x["estado"] == "CONFORME")
            agg[c] = {"estado": "CONFORME", "txt": f"{c.title()} conforme en {n} batch(es)"}
        elif any(x["estado"] == "NO APLICA" for x in e):
            na = [x for x in e if x["estado"] == "NO APLICA"]
            agg[c] = {"estado": "NO APLICA", "txt": "; ".join(sorted({x["txt"] for x in na}))}
        else:
            agg[c] = {"estado": "SIN DATO", "txt": ""}
    return agg, porbat


# ----------------------------------------------------------------- universo de lotes
lotes_stock = sorted(stock["_L"].unique())
lotes_det = sorted(det.loc[det["_VIGENTE"], "_L"].unique())

# una detencion puede apuntar a un lote que aun no llega a bodega (lag de informacion)
huerfanos = [d for d in lotes_det if not any(emparenta(d, s) for s in lotes_stock)]

# Lo mismo por el lado del laboratorio: un lote puede tener resultado NO CONFORME y todavia
# no figurar en ningun stock (produccion reciente, o ya despachada). Si no se incorpora al
# universo queda invisible, que es el peor error posible en un control de bloqueos.
lab_solo = []
for _l, _g in lab.groupby("_L"):
    if len(_l) < MIN_LOTE:
        continue
    if any(emparenta(_l, s) for s in lotes_stock) or any(emparenta(_l, d) for d in lotes_det):
        continue
    _a, _ = resume(_g)
    if any(_a[c]["estado"] == "NO CONFORME" for c in CRIT):
        lab_solo.append(_l)

universo = lotes_stock + huerfanos + lab_solo
print(f"\nUniverso: {len(lotes_stock)} lotes en stock + {len(huerfanos)} solo con detencion "
      f"+ {len(lab_solo)} solo en laboratorio y no conformes")

# Posibles typos de @ / B: calzarian si se ignorara el prefijo.
# No se marca cuando ambas variantes tienen respaldo propio en lab o en el registro:
# ahi la diferencia de prefijo es deliberada (2AS2621170M vs B2AS2621170M son lotes reales
# y distintos, listados por separado en el correo del 09/07/2026).
conocidos = set(lab["_L"]) | set(det["_L"])


def respaldado(x):
    return any(emparenta(k, x) for k in conocidos)


casi = {}
for d in huerfanos:
    cand = [s for s in lotes_stock
            if emparenta(laxo(d), laxo(s)) and not (respaldado(d) and respaldado(s))]
    if cand:
        casi[d] = cand

# Convenciones de prefijo: @ = ASC, B = BAP. Configuran lotes distintos y se respetan.
# Un par @X / BX es legitimo (mismo correlativo, dos certificaciones). Un par X / @X o
# X / BX no deberia existir: a uno de los dos le falta el prefijo.
def bare(x):
    return re.sub(r"^[@B]", "", x)


_grp = {}
for s in lotes_stock:
    _grp.setdefault(bare(s), []).append(s)
gemelos, sospecha_pfx = {}, {}
for _b, g in _grp.items():
    if len(g) < 2:
        continue
    destino = sospecha_pfx if any(x[0] not in "@B" for x in g) else gemelos
    for x in g:
        destino[x] = [y for y in g if y != x]

# Confusion de caracteres dentro del codigo (letra O contra cero, I contra uno).
_conf = {}
for s in lotes_stock:
    _conf.setdefault(s.replace("O", "0").replace("I", "1"), []).append(s)
confusion = {x: [y for y in g if y != x] for g in _conf.values() if len(g) > 1 for x in g}

# Codigos de lote que no parecen codigos: placeholders, pruebas, texto libre.
def sospechoso(txt):
    t = str(txt).strip()
    return bool(re.search(r"\s", t) or re.search(r"PRUEBA|TEST", t, re.I) or not re.search(r"\d", t))

filas = []
for l in universo:
    en_stock = l in lotes_stock
    s = stock[stock["_L"] == l] if en_stock else stock.iloc[0:0]
    rows_lab = lab[lab["_L"].map(lambda x: emparenta(x, l))]
    dets = det[det["_L"].map(lambda x: emparenta(x, l))]
    dets_v = dets[dets["_VIGENTE"]]

    # ---- batches de ahumado: la letra final del lote es parte del codigo y configura
    # lote distinto. El lab y los correos trabajan a ese nivel; el stock solo registra el
    # lote base, asi que un lote de stock agrupa varios batches que no se pueden separar.
    agg, porbat = resume(rows_lab)
    batches = {b: [porbat[b][c]["txt"] for c in CRIT
                   if porbat[b][c]["estado"] == "NO CONFORME"] for b in porbat}
    malos_bat = {b: v for b, v in batches.items() if v}
    cand_bat = {b: [porbat[b][c]["txt"] for c in CRIT
                    if porbat[b][c]["estado"] == "REMUESTREO CONFORME"]
                for b in porbat if not batches[b]}
    cand_bat = {b: v for b, v in cand_bat.items() if v}

    causas = [c for c in CRIT if agg[c]["estado"] == "NO CONFORME"]
    causas_rem = [c for c in CRIT if agg[c]["estado"] == "REMUESTREO CONFORME"]
    motivos_lab = [agg[c]["txt"] for c in causas]
    motivos_rem = [agg[c]["txt"] for c in causas_rem]
    r, _ = evalua(rows_lab)   # solo para los valores que se muestran en columnas

    # Trazabilidad de la liberacion: por que quedo liberado, criterio por criterio.
    por_crit = "; ".join(f"{c}={agg[c]['estado']}" for c in CRIT)
    _con = mide(rows_lab, "LISTERIA")
    _mu = rows_lab[rows_lab["_LM_DATO"] | rows_lab["_RAM_MAX"].notna() | rows_lab["_NIT"].notna()]
    evidencia = ", ".join(
        f"lab {cod_lab(x['CÓDIGO LAB'])} del {x['_FECHA']:%d/%m/%Y}"
        for _, x in _mu.sort_values("_FECHA").tail(4).iterrows()
        if pd.notna(x.get("CÓDIGO LAB")) and pd.notna(x["_FECHA"]))
    motivo_lib = [agg[c]["txt"] for c in CRIT
                  if agg[c]["estado"] in ("CONFORME", "NO APLICA") and agg[c]["txt"]]

    # ---- detenciones vigentes y su evaluabilidad
    motivos_det, propuestas = [], []
    hay_pnc = (dets_v["_ESTADO"] == "PNC").any()
    for _, d in dets_v.iterrows():
        etq = f"{d['ID']} {d['TIPO DE DESVIACION']}"
        if d["_ESTADO"] == "PNC":
            motivos_det.append(f"{etq} (PNC)")
            propuestas.append(f"{d['ID']}: PNC - no se libera contra laboratorio")
            continue
        motivos_det.append(etq)
        crit = [c.strip().upper() for c in str(d["CRITERIOS DE LIBERACION"]).split(";") if c.strip()]
        rp, post = evalua(rows_lab, desde=d["_EVENTO"])
        if rp["n"] == 0:
            propuestas.append(f"{d['ID']}: NO LIBERABLE - sin muestras posteriores al "
                              f"{d['_EVENTO']:%d/%m/%Y}")
            continue
        malos = [txt for c in crit for mal, _f, txt in [incumple(rp, c)] if mal]
        faltan = [c for c in crit if incumple(rp, c)[1]]
        if malos:
            propuestas.append(f"{d['ID']}: NO LIBERABLE - resultado no conforme ({'; '.join(malos)})")
        elif faltan:
            propuestas.append(f"{d['ID']}: NO LIBERABLE - falta {', '.join(faltan)} "
                              f"en las {rp['n']} muestras posteriores al evento")
        else:
            propuestas.append(f"{d['ID']}: LIBERABLE segun lab - {rp['n']} muestra(s) posterior(es) "
                              f"conforme(s) en {', '.join(crit)}. Requiere firma de Calidad.")

    # ---- estado consolidado
    if hay_pnc:
        estado, origen = "PNC", "DETENCION"
    elif motivos_lab and motivos_det:
        estado, origen = "BLOQUEADO", "LAB + DETENCION"
    elif motivos_lab:
        estado, origen = "BLOQUEADO", "LAB"
    elif motivos_det:
        estado, origen = "BLOQUEADO", "DETENCION"
    elif motivos_rem:
        # fallo y se re-muestreo conforme cubriendo el criterio: no se libera solo
        estado, origen = "CANDIDATO A LIBERAR", "LAB"
    elif len(rows_lab) == 0 or all(agg[c]["estado"] == "SIN DATO" for c in CRIT):
        estado, origen = "SIN ANALISIS", ""
    else:
        estado, origen = "LIBERADO", ""

    obs = []
    if not en_stock and len(dets):
        obs.append("Lote con detencion vigente que NO aparece en ningun stock: "
                   "puede ser lag de informacion, producto aun en proceso, ya despachado, "
                   "bodega no incluida, o lote mal escrito. Requiere revision.")
    elif not en_stock:
        obs.append("Lote con resultado NO CONFORME que no aparece en ningun stock: "
                   "produccion reciente aun no ingresada a bodega, ya despachada, o bodega no "
                   "incluida. Verificar antes de despachar.")
    if l in casi:
        obs.append("Posible error de tipeo en prefijo @/B: se parece a " + ", ".join(casi[l]))
    if l in sospecha_pfx:
        obs.append("SOSPECHA DE TIPEO: convive en stock con " + ", ".join(sospecha_pfx[l]) +
                   ". @ (ASC) y B (BAP) configuran lotes distintos, pero un lote sin prefijo "
                   "junto a su gemelo prefijado no deberia existir: a uno le falta el prefijo")
    if l in gemelos:
        obs.append("Convive en stock con " + ", ".join(gemelos[l]) +
                   " (ASC vs BAP): son lotes distintos, sin accion")
    if l in confusion:
        obs.append("Se confunde con " + ", ".join(confusion[l]) + " por letra O contra cero")
    # manda la clasificacion de la muestra: el laboratorio nombra la linea, la
    # condicion de bodega solo dice como esta guardado hoy
    _lm = sorted({x for x in rows_lab["_LINEA"] if x})
    _ls = linea_stock(l)
    linea_l = " / ".join(_lm) if _lm else (_ls or "sin clasificar")
    if _lm and _ls and _ls not in _lm:
        obs.append(f"El laboratorio clasifica este lote como {' / '.join(_lm)} y en bodega "
                   f"figura como {_ls}: manda el laboratorio para aplicar el nitrito")
    if rows_lab["_BW"].any() and (rows_lab["_LINEA"] == "CONGELADA").any():
        obs.append("Bacon/wheel: figura congelado pero se vende refrigerado en destino, "
                   "asi que el nitrito se evalua igual que en linea refrigerada")
    ign = rows_lab[~rows_lab["_NIT_APLICA"] & rows_lab["_NIT"].notna()
                   & (rows_lab["_NIT"] < LIM_NITRITO)]
    if len(ign):
        obs.append(f"Nitrito bajo ({ign['_NIT'].min():.1f} ppm) NO aplicado: linea congelada")
    lis_ex = rows_lab[rows_lab["_LM_P"] & ~rows_lab["_LIS_APLICA"]]
    if len(lis_ex):
        obs.append("Listeria PRESENCIA NO aplicada: linea congelada sin destino EE.UU. "
                   "ni Costa Rica (cliente " + " / ".join(sorted(set(
                       s["CLIENTE"].dropna().astype(str)))[:2] or ["sin dato"]) + ")")
    lib_f, lib_m, lib_o = liberacion(l)
    if lib_f and (motivos_lab or motivos_det):
        obs.append(f"CONFLICTO: figura liberado el {lib_f}"
                   + (f" para {lib_m}" if lib_m else "")
                   + f" en {lib_o}, pero el laboratorio mantiene un incumplimiento vigente")
    elif lib_f and motivos_rem:
        obs.append(f"Ya figura liberado el {lib_f}" + (f" para {lib_m}" if lib_m else "")
                   + f" en {lib_o}: el re-muestreo conforme respalda esa decision")
    if cand_bat:
        obs.append(f"Re-muestreo conforme posterior en {len(cand_bat)} batch(es): "
                   + " | ".join(t for v in cand_bat.values() for t in v)
                   + ". Requiere decision firmada de Calidad, el sistema no libera solo")
    if malos_bat and len(malos_bat) < len(batches):
        obs.append(f"Bloqueo originado en {len(malos_bat)} de {len(batches)} batches de ahumado "
                   "con resultado. El stock no registra la letra de batch, asi que las cajas de "
                   "los batches conformes no se pueden separar y quedan bloqueadas con el lote")
    if en_stock:
        etiqueta = s["LOTE DE PLANTA"].iloc[0]
    elif len(dets):
        etiqueta = dets["LOTE"].iloc[0]
    else:
        etiqueta = str(rows_lab["LOTE SW"].iloc[0]).strip().split("*")[0] if len(rows_lab) else l
    if sospechoso(etiqueta):
        obs.append("El codigo de lote no parece un lote real (placeholder o texto libre): "
                   "revisar el origen en Fishken, agrupa cajas que no comparten lote")
    if len(rows_lab) == 0 and en_stock:
        obs.append("Sin resultados en LAB-REG-08 2026")
    else:
        faltan = [c for c in ("LISTERIA", "RAM") if agg[c]["estado"] == "SIN DATO"]
        if faltan:
            obs.append("Sin analisis de " + " y ".join(faltan))
    if r["ar"] == "R" and not any("Nitrito" in m for m in motivos_lab):
        obs.append("Lab marco nitrito RECHAZADO (criterio propio del producto)")
    tipos = set(rows_lab["TIPO "].dropna().astype(str).str.strip())
    if tipos - {"PT"}:
        obs.append("Incluye muestras tipo " + "/".join(sorted(tipos)))
    if (dets_v["ALCANCE"].astype(str).str.startswith("PARCIAL")).any():
        obs.append("Detencion de alcance PARCIAL: sin identificar cajas se marca el lote completo")

    filas.append({
        "LOTE": etiqueta,
        "LOTE (normalizado)": l,
        "ESTADO": estado,
        "ORIGEN DEL BLOQUEO": origen,
        "MOTIVO DE LA LIBERACION": (" | ".join(motivo_lib)
                                    if estado in ("LIBERADO", "CANDIDATO A LIBERAR") else ""),
        "ESTADO POR CRITERIO": por_crit,
        "EVIDENCIA (ultimas muestras)": evidencia,
        "LINEA": linea_l,
        "DESTINO RESTRINGIDO": (lambda d: d if d else
                                ("sin determinar" if d is None else "no"))(destino_restringido(l)),
        "CAUSAS LAB": ",".join(causas),
        "CAUSAS CON REMUESTREO CONFORME": ",".join(causas_rem),
        "LIBERACION DECLARADA": lib_f,
        "MERCADOS LIBERADOS": lib_m,
        "FUENTE DE LA LIBERACION": lib_o,
        "HISTORIA DEL RE-MUESTREO": " | ".join(motivos_rem),
        "TIPOS DE DESVIACION": ",".join(sorted(
            dets_v["TIPO DE DESVIACION"].dropna().astype(str).str.strip().unique())),
        "MOTIVO - LABORATORIO": " | ".join(motivos_lab),
        "MOTIVO - DETENCION": " | ".join(motivos_det),
        "PROPUESTA DE LIBERACION (no libera)": " || ".join(propuestas),
        "OBSERVACIONES": " | ".join(obs),
        "LISTERIA": r["listeria"] or ("sin dato" if len(rows_lab) else None),
        "RAM MAX (UFC/g)": r["ram"],
        "NITRITO PROM. MIN (ppm)": r["nitrito"],
        "A/R NITRITO (lab)": r["ar"],
        "BATCHES CON RESULTADO": len(batches),
        "BATCHES NO CONFORMES": "; ".join(f"{b[len(l):] or '(base)'}: {', '.join(v)}"
                                          for b, v in malos_bat.items()),
        "N MUESTRAS LAB": len(rows_lab),
        "N DETENCIONES VIGENTES": len(dets_v),
        "ULTIMA MUESTRA LAB": rows_lab["_FECHA"].max() if len(rows_lab) else pd.NaT,
        "EN STOCK": "SI" if en_stock else "NO",
        "BODEGA(S)": " / ".join(sorted(s["BODEGA"].dropna().astype(str).unique())) if en_stock else "",
        "PRODUCTOS": (" / ".join(sorted(s["NOMBRE PRODUCTO"].dropna().astype(str).unique())) if en_stock
                      else " / ".join(sorted(dets["PRODUCTO"].dropna().astype(str).unique())) if len(dets)
                      else " / ".join(sorted(rows_lab["PRESENTACIÓN"].dropna().astype(str).unique()))),
        "CLIENTE(S)": " / ".join(sorted(s["CLIENTE"].dropna().astype(str).unique())) if en_stock else "",
        "CAJAS EN STOCK": len(s),
    })

res = pd.DataFrame(filas)
orden = {"PNC": 0, "BLOQUEADO": 1, "CANDIDATO A LIBERAR": 2, "SIN ANALISIS": 3, "LIBERADO": 4}
res = res.sort_values(["ESTADO", "EN STOCK", "LOTE"],
                      key=lambda c: c.map(orden) if c.name == "ESTADO" else c).reset_index(drop=True)

# ------------------------------------------------- veredicto por batch (lab)
# Son DOS preguntas distintas contra la misma fuente, y no deben mezclarse:
#   packing list -> viene con el batch exacto: se resuelve contra LAB-REG-08 y punto.
#   stock        -> solo trae el lote base: ahi si hay que agregar los batches, con
#                   la perdida de precision que eso implica.
# Esta hoja responde la primera. La hoja RESUMEN POR LOTE responde la segunda.
bat = []
for b, g in lab.groupby("_L"):
    rb, _ = evalua(g)
    hb = {c: historia(g, c) for c in CRIT}
    causas_b = [c for c in CRIT if hb[c]["estado"] == "NO CONFORME"]
    rem_b = [c for c in CRIT if hb[c]["estado"] == "REMUESTREO CONFORME"]
    dv = det[det["_L"].map(lambda x: emparenta(x, b)) & det["_VIGENTE"]]
    base = next((s for s in lotes_stock if emparenta(s, b)), "")
    juzgable = any(hb[c]["estado"] in ("CONFORME", "NO CONFORME", "REMUESTREO CONFORME")
                   for c in CRIT)
    if (dv["_ESTADO"] == "PNC").any():
        est = "PNC"
    elif causas_b or len(dv):
        est = "BLOQUEADO"
    elif rem_b:
        est = "CANDIDATO A LIBERAR"
    elif juzgable:
        est = "LIBERADO"
    else:
        est = "SIN ANALISIS"
    lb_f, lb_m, lb_o = liberacion(b)
    bat.append({
        "BATCH": str(g["LOTE SW"].iloc[0]).strip().split("*")[0],
        "BATCH (normalizado)": b,
        "ESTADO": est,
        "CAUSAS LAB": ",".join(causas_b),
        "MOTIVO - LABORATORIO": " | ".join(hb[c]["txt"] for c in causas_b),
        "RE-MUESTREO CONFORME": " | ".join(hb[c]["txt"] for c in rem_b),
        "MOTIVO - DETENCION": " | ".join(f"{d['ID']} {d['TIPO DE DESVIACION']}"
                                         for _, d in dv.iterrows()),
        "LISTERIA": rb["listeria"] or "sin dato",
        "RAM MAX (UFC/g)": rb["ram"],
        "NITRITO PROMEDIO (ppm)": rb["nitrito"],
        "N MUESTRAS": rb["n"],
        "PRIMERA MUESTRA": g["_FECHA"].min(),
        "ULTIMA MUESTRA": g["_FECHA"].max(),
        "TIPO": " / ".join(sorted(g["TIPO "].dropna().astype(str).str.strip().unique())),
        "PRESENTACION": " / ".join(sorted(g["PRESENTACIÓN"].dropna().astype(str).unique()))[:120],
        "OBSERVACION LAB": " / ".join(sorted(g["OBSERVACIÓN"].dropna().astype(str).unique()))[:120],
        "LINEA": " / ".join(sorted({x for x in g["_LINEA"] if x})) or "sin clasificar",
        "LIBERACION DECLARADA": lb_f,
        "MERCADOS LIBERADOS": lb_m,
        "FUENTE DE LA LIBERACION": lb_o,
        "LOTE BASE EN STOCK": base,
    })
bats = pd.DataFrame(bat).sort_values("BATCH (normalizado)")
print(f"Veredicto por batch: {len(bats)} batches con resultado | "
      + " ".join(f"{k}={v}" for k, v in bats['ESTADO'].value_counts().items()))

# ------------------------------------------------- historial de veredictos
# El cruce se recalcula entero en cada corrida: sin esto la historia se pierde y
# no hay forma de saber que un lote estuvo bloqueado y despues se libero.
# estado_actual.csv se sobreescribe (es la foto contra la que se compara);
# cambios.csv solo crece y es el registro de transiciones.
HIST = config.ruta(config.DIR_HISTORIAL)
os.makedirs(HIST, exist_ok=True)
F_ACT = os.path.join(HIST, "estado_actual.csv")
F_CAM = os.path.join(HIST, "cambios.csv")
CORRIDA = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
# Huella de los criterios vigentes: sin esto no hay forma de saber si un lote
# cambio de estado porque llego un resultado nuevo o porque cambiamos la regla.
CRITERIOS = config.huella_criterios()
F_CRI = os.path.join(HIST, "criterios.csv")
_prev_cri = ""
if os.path.exists(F_CRI):
    _c = pd.read_csv(F_CRI, dtype=str).fillna("")
    _prev_cri = _c["CRITERIOS"].iloc[-1] if len(_c) else ""
CAMBIO_CRI = "SI" if (_prev_cri and _prev_cri != CRITERIOS) else "NO"
if _prev_cri != CRITERIOS:
    pd.DataFrame([[CORRIDA, CRITERIOS]], columns=["CORRIDA", "CRITERIOS"]).to_csv(
        F_CRI, mode="a", header=not os.path.exists(F_CRI), index=False, encoding="utf-8-sig")
    print(f"  criterios: {'CAMBIARON en esta corrida' if _prev_cri else 'registrados'}")

ahora = pd.concat([
    res.assign(NIVEL="lote")[["NIVEL", "LOTE (normalizado)", "ESTADO", "CAUSAS LAB",
                              "CAJAS EN STOCK"]]
       .rename(columns={"LOTE (normalizado)": "CLAVE", "CAJAS EN STOCK": "CAJAS"}),
    bats.assign(NIVEL="batch", CAJAS=0)[["NIVEL", "BATCH (normalizado)", "ESTADO",
                                         "CAUSAS LAB", "CAJAS"]]
        .rename(columns={"BATCH (normalizado)": "CLAVE"}),
], ignore_index=True)
ahora["CAUSAS LAB"] = ahora["CAUSAS LAB"].fillna("")
ahora["CAJAS"] = ahora["CAJAS"].fillna(0).astype(int)

prev = None
if os.path.exists(F_ACT):
    _a = pd.read_csv(F_ACT, dtype=str).fillna("")
    prev = dict(zip(_a["NIVEL"] + "|" + _a["CLAVE"], _a["ESTADO"]))

cambios = []
if prev is not None:
    for _, x in ahora.iterrows():
        k = f"{x['NIVEL']}|{x['CLAVE']}"
        ant = prev.get(k)
        if ant is None:
            cambios.append([CORRIDA, x["NIVEL"], x["CLAVE"], "(nuevo)", x["ESTADO"],
                            x["CAUSAS LAB"], x["CAJAS"], CAMBIO_CRI])
        elif ant != x["ESTADO"]:
            cambios.append([CORRIDA, x["NIVEL"], x["CLAVE"], ant, x["ESTADO"],
                            x["CAUSAS LAB"], x["CAJAS"], CAMBIO_CRI])
    vivos = set(ahora["NIVEL"] + "|" + ahora["CLAVE"])
    for k, v in prev.items():
        if k not in vivos:
            nv, cl = k.split("|", 1)
            cambios.append([CORRIDA, nv, cl, v, "(ya no aparece)", "", 0, CAMBIO_CRI])

COLS_CAM = ["CORRIDA", "NIVEL", "CLAVE", "ESTADO ANTERIOR", "ESTADO NUEVO", "CAUSAS", "CAJAS",
            "HUBO CAMBIO DE CRITERIO"]


def leer_cambios():
    """Lee el log tolerando que haya crecido de esquema entre versiones.

    Antes se hacia append y al agregar una columna el archivo quedo con filas de
    7 y de 8 campos, que el parser rapido no puede leer. Ahora se reescribe
    completo en cada corrida (son cientos de filas) y ademas se normaliza lo viejo.
    """
    if not os.path.exists(F_CAM):
        return pd.DataFrame(columns=COLS_CAM)
    for kw in ({}, {"engine": "python", "on_bad_lines": "skip"}):
        try:
            d = pd.read_csv(F_CAM, dtype=str, **kw).fillna("")
            break
        except Exception:                                        # noqa: BLE001
            d = None
    if d is None:
        bak = F_CAM + ".bak"
        os.replace(F_CAM, bak)
        print(f"  historial ilegible, se movio a {os.path.basename(bak)} y se empieza de nuevo")
        return pd.DataFrame(columns=COLS_CAM)
    for c in COLS_CAM:
        if c not in d.columns:
            d[c] = ""
    return d[d["CORRIDA"] != "CORRIDA"][COLS_CAM]


camlog = pd.concat([leer_cambios(), pd.DataFrame(cambios, columns=COLS_CAM)], ignore_index=True)
camlog.to_csv(F_CAM, index=False, encoding="utf-8-sig")
ahora.to_csv(F_ACT, index=False, encoding="utf-8-sig")

if prev is None:
    print("\nHistorial: linea base creada. Desde la proxima corrida se registran los cambios.")
else:
    print(f"\nHistorial: {len(cambios)} cambio(s) de estado desde la corrida anterior")
    for c in cambios[:12]:
        print(f"   {c[1]:5s} {c[2]:16s} {c[3]}  ->  {c[4]}"
              + (f"  ({c[6]} cajas)" if c[6] else ""))
    if len(cambios) > 12:
        print(f"   ... y {len(cambios) - 12} mas")

camlog = camlog.iloc[::-1]          # lo mas reciente primero

# ----------------------------------------------------------------- detalles
mapa = res.set_index("LOTE (normalizado)")
d_st = stock.copy()
for col, src in (("ESTADO", "ESTADO"), ("ORIGEN DEL BLOQUEO", "ORIGEN DEL BLOQUEO"),
                 ("MOTIVO - LABORATORIO", "MOTIVO - LABORATORIO"),
                 ("MOTIVO - DETENCION", "MOTIVO - DETENCION")):
    d_st[col] = d_st["_L"].map(mapa[src])
cols_st = (["ESTADO", "ORIGEN DEL BLOQUEO", "MOTIVO - LABORATORIO", "MOTIVO - DETENCION", "BODEGA"]
           + [c for c in stock.columns if not c.startswith("_") and c != "BODEGA"])
d_st = d_st[cols_st]

lab["_LOTE_REF"] = None
for l in universo:
    m = lab["_L"].map(lambda x: emparenta(x, l))
    for i in lab.index[m]:
        p = lab.at[i, "_LOTE_REF"]
        if p is None or len(l) > len(p):
            lab.at[i, "_LOTE_REF"] = l
d_lab = lab[lab["_LOTE_REF"].notna()].copy()
d_lab["ESTADO LOTE"] = d_lab["_LOTE_REF"].map(mapa["ESTADO"])
d_lab = d_lab[["_LOTE_REF", "ESTADO LOTE", "FUENTE LAB", "CÓDIGO LAB", "FECHA INGRESO", "TIPO ",
               "GRUPO", "PRESENTACIÓN", "LOTE SW", "OBSERVACIÓN"] + LM + RAM + NIT +
              ["NITRITO PROMEDIO", "A/R Nitrito"]].rename(
    columns={"_LOTE_REF": "LOTE (normalizado)"})
d_lab = d_lab.sort_values(["LOTE (normalizado)", "FECHA INGRESO"])

d_det = det.drop(columns=[c for c in det.columns if c.startswith("_")]).copy()
d_det["ESTADO DEL LOTE EN EL CRUCE"] = det["_L"].map(
    lambda x: next((mapa.at[l, "ESTADO"] for l in universo if emparenta(x, l)), "sin match"))
d_det["PROPUESTA DE LIBERACION"] = det["_L"].map(
    lambda x: next((mapa.at[l, "PROPUESTA DE LIBERACION (no libera)"]
                    for l in universo if emparenta(x, l)), ""))

# ----------------------------------------------------------------- escritura
# Ventana de cada fuente: es lo que explica el desfase entre bodega, laboratorio y correos.
fuentes = []
for a, g in stock.groupby("ARCHIVO ORIGEN"):
    f = pd.to_datetime(g["FECHA INGRESO"], errors="coerce")
    fuentes.append({"FUENTE": "Stock", "ARCHIVO": a, "REGISTROS": len(g),
                    "DESDE": f.min(), "HASTA": f.max()})
for h, g in lab.groupby("FUENTE LAB"):
    fuentes.append({"FUENTE": "Laboratorio", "ARCHIVO": h, "REGISTROS": len(g),
                    "DESDE": g["_FECHA"].min(), "HASTA": g["_FECHA"].max()})
fd = pd.to_datetime(det["FECHA CORREO"], errors="coerce")
fuentes.append({"FUENTE": "Detenciones", "ARCHIVO": os.path.basename(F_DET), "REGISTROS": len(det),
                "DESDE": fd.min(), "HASTA": fd.max()})
fuentes = pd.DataFrame(fuentes)

with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    res.to_excel(xl, sheet_name="RESUMEN POR LOTE", index=False, startrow=3)
    bats.to_excel(xl, sheet_name="VEREDICTO POR BATCH", index=False)
    camlog.to_excel(xl, sheet_name="CAMBIOS", index=False)
    fuentes.to_excel(xl, sheet_name="FUENTES", index=False)
    d_st.to_excel(xl, sheet_name="DETALLE STOCK", index=False)
    d_lab.to_excel(xl, sheet_name="DETALLE LAB", index=False)
    d_det.to_excel(xl, sheet_name="DETENCIONES", index=False)

wb = load_workbook(OUT)
ws, wd = wb["RESUMEN POR LOTE"], wb["DETALLE STOCK"]
c_l = get_column_letter(list(d_st.columns).index("LOTE DE PLANTA") + 1)
c_kg = get_column_letter(list(d_st.columns).index("PESO NETO (KG)") + 1)
c_pz = get_column_letter(list(d_st.columns).index("PIEZAS") + 1)
nf = len(d_st) + 1
c_lr = get_column_letter(list(res.columns).index("LOTE") + 1)
k1, k2 = len(res.columns) + 1, len(res.columns) + 2
ws.cell(4, k1, "KG NETOS EN STOCK")
ws.cell(4, k2, "PIEZAS EN STOCK")
for i in range(len(res)):
    f = 5 + i
    rl = f"'DETALLE STOCK'!${c_l}$2:${c_l}${nf}"
    ws.cell(f, k1, f"=SUMIFS('DETALLE STOCK'!${c_kg}$2:${c_kg}${nf},{rl},${c_lr}{f})")
    ws.cell(f, k2, f"=SUMIFS('DETALLE STOCK'!${c_pz}$2:${c_pz}${nf},{rl},${c_lr}{f})")

ws["A1"] = "CRUCE DE BLOQUEOS - STOCK x LAB-REG-08 2026 x REGISTRO DE DETENCIONES"
ws["A2"] = (f"Laboratorio (derivado, se recalcula): Listeria PRESENCIA en toda linea | "
            f"RAM > {LIM_RAM:,} UFC/g en toda linea | Nitrito promedio < {LIM_NITRITO} ppm "
            "SOLO en linea refrigerada y en bacon/wheel, que salen congelados pero se venden "
            "refrigerados en destino (el resto de la congelada, incluidos los carpaccios, "
            "no lo aplica) | "
            "Listeria SOLO en linea refrigerada y en congelada con destino EE.UU. o "
            "Costa Rica (cliente con PMT); si el destino no se puede determinar, se aplica. "
            "Un criterio deja de estar vigente si hay re-muestreo posterior conforme que vuelva "
            "a medirlo. Detencion (declarada por correo): bloquea aunque no haya resultado.")
ws["A3"] = ("Los dos origenes se acumulan: un lote con ambos debe cerrar los dos. La columna de propuesta "
            "NO libera nada: exige muestra posterior a la fecha del evento y que cubra el criterio "
            "desviado; la liberacion la firma Calidad en el REGISTRO DETENCIONES.")
ws["A1"].font = Font(name="Arial", size=13, bold=True)
for r in (2, 3):
    ws[f"A{r}"].font = Font(name="Arial", size=9, italic=True)
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=k2)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=k2)

hf = PatternFill("solid", fgColor="1F3864")
thin = Side(style="thin", color="BFBFBF")
fills = {"PNC": PatternFill("solid", fgColor="D9D9D9"),
         "BLOQUEADO": PatternFill("solid", fgColor="FFC7CE"),
         "CANDIDATO A LIBERAR": PatternFill("solid", fgColor="DDEBF7"),
         "SIN ANALISIS": PatternFill("solid", fgColor="FFEB9C"),
         "LIBERADO": PatternFill("solid", fgColor="C6EFCE")}
fonts = {"PNC": Font(name="Arial", size=10, bold=True, color="3F3F3F"),
         "BLOQUEADO": Font(name="Arial", size=10, bold=True, color="9C0006"),
         "CANDIDATO A LIBERAR": Font(name="Arial", size=10, bold=True, color="1F4E79"),
         "SIN ANALISIS": Font(name="Arial", size=10, bold=True, color="9C6500"),
         "LIBERADO": Font(name="Arial", size=10, bold=True, color="006100")}

for hoja, h in (("RESUMEN POR LOTE", 4), ("VEREDICTO POR BATCH", 1), ("CAMBIOS", 1),
                ("DETALLE STOCK", 1), ("DETALLE LAB", 1), ("DETENCIONES", 1), ("FUENTES", 1)):
    w = wb[hoja]
    for c in w[h]:
        if c.value is not None:
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = hf
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            c.border = Border(bottom=thin)
    w.row_dimensions[h].height = 30
    w.freeze_panes = w.cell(h + 1, 3)
    w.auto_filter.ref = f"A{h}:{get_column_letter(w.max_column)}{w.max_row}"
    for col in range(1, w.max_column + 1):
        L = [len(str(w.cell(r, col).value or "")) for r in range(h, min(w.max_row, h + 300) + 1)]
        w.column_dimensions[get_column_letter(col)].width = min(max(10, max(L) + 2), 46)
    if hoja in ("RESUMEN POR LOTE", "DETALLE STOCK", "VEREDICTO POR BATCH"):
        ce = [c.column for c in w[h] if c.value == "ESTADO"][0]
        for r in range(h + 1, w.max_row + 1):
            c = w.cell(r, ce)
            if c.value in fills:
                c.fill, c.font = fills[c.value], fonts[c.value]

for nom, fmt in (("RAM MAX (UFC/g)", "#,##0"), ("NITRITO PROM. MIN (ppm)", "0.0"),
                 ("KG NETOS EN STOCK", "#,##0.00"), ("PIEZAS EN STOCK", "#,##0"),
                 ("CAJAS EN STOCK", "#,##0")):
    col = [c.column for c in ws[4] if c.value == nom][0]
    for r in range(5, ws.max_row + 1):
        ws.cell(r, col).number_format = fmt
        ws.cell(r, col).font = Font(name="Arial", size=10)

wb.save(OUT)

# ----------------------------------------------------------------- consola
print("\nLotes evaluados:", len(res), "| en stock:", len(lotes_stock), "| solo detencion:", len(huerfanos))
print(res["ESTADO"].value_counts().to_string())
print("\n--- BLOQUEADOS / PNC ---")
v = res[res["ESTADO"].isin(["BLOQUEADO", "PNC"])]
print(v[["LOTE", "ESTADO", "ORIGEN DEL BLOQUEO", "MOTIVO - LABORATORIO", "MOTIVO - DETENCION",
         "EN STOCK", "CAJAS EN STOCK"]].to_string(index=False))
print("\n--- PROPUESTAS DE LIBERACION ---")
for _, x in res[res["PROPUESTA DE LIBERACION (no libera)"] != ""].iterrows():
    print(f"  {x['LOTE']:14s} {x['PROPUESTA DE LIBERACION (no libera)']}")
print("\n--- CALIDAD DE DATOS ---")
print("Sospecha de prefijo faltante:", {k: v for k, v in sospecha_pfx.items()} or "ninguna")
print("Pares ASC/BAP legitimos:", sorted({tuple(sorted([k] + v)) for k, v in gemelos.items()}) or "ninguno")
print("Confusion O/0:", sorted({tuple(sorted([k] + v)) for k, v in confusion.items()}) or "ninguna")
print("Detenciones sin lote en stock:", casi or "sin sospecha de tipeo")

sobre = res[(res["BATCHES NO CONFORMES"] != "") &
            (res["ESTADO"].isin(["BLOQUEADO", "PNC"]))].copy()
sobre["_nb"] = sobre["BATCHES NO CONFORMES"].str.count(";") + 1
parcial = sobre[sobre["_nb"] < sobre["BATCHES CON RESULTADO"]]
print(f"\nLotes bloqueados por un subconjunto de sus batches: {len(parcial)} "
      f"({int(parcial['CAJAS EN STOCK'].sum())} cajas) - no separables con el dato actual")
print("Salida:", OUT)
