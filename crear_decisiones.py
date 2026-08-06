# -*- coding: utf-8 -*-
"""Crea REGISTRO DECISIONES.xlsx (solo si no existe).

Es el registro de las liberaciones firmadas. Junto con REGISTRO DETENCIONES es
lo DECLARADO del sistema: lo escriben personas, el motor solo lo lee y nunca lo
sobrescribe. Sus columnas son las mismas de la tabla `decision` de sql/, para
que pasar a base de datos sea copiar y no rediseniar.
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config

OUT = config.ruta("REGISTRO DECISIONES.xlsx")
if os.path.exists(OUT):
    raise SystemExit("Ya existe " + OUT + " - no se sobrescribe.")

COLS = ["ID", "FECHA", "TIPO", "LOTE", "BATCH", "DETENCION", "ANULA",
        "MERCADOS", "EVIDENCIA", "HUELLA DE CRITERIOS AL FIRMAR", "COMENTARIO",
        "FIRMADO POR", "RUT"]

EJEMPLO = ["EJEMPLO-001", "2026-08-06", "LIBERACION", "@2CK2605028", "", "",
           "", "Nacional/Exportacion",
           "lab 402 del 11/02/2026 conforme en LISTERIA tras altas presiones",
           config.huella_criterios(),
           "Re-muestreo posterior al hallazgo cubre el criterio que fallo",
           "Nombre Apellido - Jefe de Calidad", "12345678-9"]

GUIA = [
    ("ID", "Correlativo DEC-AAAA-NNN. Las filas que empiezan con EJEMPLO se ignoran."),
    ("FECHA", "Cuando se firma. Debe ser posterior a la muestra que sirve de evidencia."),
    ("TIPO", "LIBERACION libera un lote o batch | CIERRE DETENCION cierra una detencion "
             "| ANULACION deja sin efecto una decision anterior."),
    ("LOTE", "Lote de planta con su prefijo @ o B. Libera el lote completo."),
    ("BATCH", "Opcional y mas preciso: libera solo ese batch de ahumado. Si se llena, "
              "manda sobre LOTE."),
    ("DETENCION", "ID de la detencion que se cierra. Solo para CIERRE DETENCION."),
    ("ANULA", "ID de la decision que queda sin efecto. Solo para ANULACION. "
              "Una decision NUNCA se borra ni se edita: se anula con una nueva."),
    ("MERCADOS", "Para que mercados se libera, separados por /. Vacio = todos. "
                 "Ojo: EE.UU. y Costa Rica exigen ausencia de Listeria."),
    ("EVIDENCIA", "Contra que se libera: codigo y fecha de las muestras. Sin esto la "
                  "decision no es auditable."),
    ("HUELLA DE CRITERIOS AL FIRMAR", "Se copia de historial/criterios.csv. Permite saber "
                                      "si las reglas cambiaron despues de la firma."),
    ("COMENTARIO", "Por que se decidio. En una linea."),
    ("FIRMADO POR", "Nombre y cargo de quien tiene la atribucion. No de quien tipea."),
    ("RUT", "RUT de quien firma, formato 12345678-9."),
]

df = pd.DataFrame([EJEMPLO], columns=COLS)
with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    df.to_excel(xl, sheet_name="DECISIONES", index=False, startrow=4)
    pd.DataFrame(GUIA, columns=["CAMPO", "QUE SE ESCRIBE"]).to_excel(
        xl, sheet_name="INSTRUCTIVO", index=False)

wb = load_workbook(OUT)
ws = wb["DECISIONES"]
ws["A1"] = "REGISTRO DE DECISIONES DE LIBERACION"
ws["A2"] = ("Lo escribe Calidad. El motor lo LEE y nunca lo sobrescribe. Una liberacion "
            "firmada aca es lo unico que convierte un CANDIDATO A LIBERAR en LIBERADO.")
ws["A3"] = ("Una firma vale para la evidencia que existia al firmarla: si despues llega un "
            "resultado no conforme, el motor vuelve a bloquear y avisa que la liberacion "
            "quedo superada. Para deshacer una decision se ANULA con otra fila, no se borra.")
ws["A1"].font = Font(name="Arial", size=13, bold=True)
for r in (2, 3):
    ws[f"A{r}"].font = Font(name="Arial", size=9, italic=True)
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
ws.row_dimensions[2].height = 26
ws.row_dimensions[3].height = 38

fill = PatternFill("solid", fgColor="1F3864")
for c in ws[5]:
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = Border(bottom=Side(style="thin", color="BFBFBF"))
ws.row_dimensions[5].height = 34
ws.freeze_panes = "C6"
ws.auto_filter.ref = f"A5:{get_column_letter(len(COLS))}{ws.max_row}"

ancho = {"ID": 14, "FECHA": 12, "TIPO": 18, "LOTE": 16, "BATCH": 18, "DETENCION": 14,
         "ANULA": 13, "MERCADOS": 20, "EVIDENCIA": 46,
         "HUELLA DE CRITERIOS AL FIRMAR": 40, "COMENTARIO": 40,
         "FIRMADO POR": 28, "RUT": 14}
for i, c in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = ancho[c]

gris = PatternFill("solid", fgColor="F2F2F2")
for r in range(6, ws.max_row + 1):
    for i, c in enumerate(COLS, 1):
        cell = ws.cell(r, i)
        cell.font = Font(name="Arial", size=10, italic=(r == 6), color="808080" if r == 6 else "000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r == 6:
            cell.fill = gris
        if c == "FECHA":
            cell.number_format = "DD/MM/YYYY"

dv = DataValidation(type="list", formula1='"LIBERACION,CIERRE DETENCION,ANULACION"',
                    allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"C6:C500")

wi = wb["INSTRUCTIVO"]
for c in wi[1]:
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = fill
wi.column_dimensions["A"].width = 32
wi.column_dimensions["B"].width = 100
for r in range(2, wi.max_row + 1):
    wi.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
    wi.cell(r, 2).font = Font(name="Arial", size=10)
    wi.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print("Creado:", OUT)
